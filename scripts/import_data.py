"""Import enrichi depuis le classeur Google Sheets (export .xlsx).

Charge les vraies données de la pisciculture dans l'application, **via l'API
HTTP** — donc à travers les invariants métier (une ligne invalide est refusée,
pas insérée en douce).

Format attendu du classeur — onglets et colonnes (indices 0-based) :
    « Chargés de projets »        : [0] nom
    « Projets »                   : [0] nom  [1] fin prévue  [2] jalon de fin max
                                    [3] raison fin  [4] trigramme projet
                                    [5] trigramme epic  [6] rappel  [7] terminé ?
    « Projet non plannifiés »     : [0] nom
    « Detail des tache de projet »: [0] trigramme projet  [2] nom
                                    [3] date début  [4] date fin
                                    [6] responsable  [9] terminé ?
    « Jalons »                    : [0] nom  [1] date

    `scripts/make_sample_source.py` génère un classeur d'exemple conforme à ce
    format — c'est la spec exécutable, et de quoi tester l'import sans données
    réelles.

Utilisation :
    pip install -e "api/[scripts]"          # requests + openpyxl
    python scripts/import_data.py \\
        --api http://localhost:8080 \\
        --xlsx data/source.xlsx \\
        --email admin@… --password …

Authentification (l'import crée des utilisateurs ⇒ droits admin requis) :
    --email/--password  : le script se connecte et récupère un jeton, ou
    --token <jwt>       : jeton fourni directement, ou
    aucun des deux      : suppose AUTH_DISABLED=true côté API (dev uniquement).

Idempotent : ré-exécutable sans créer de doublons.

Correspondances des clés étrangères :
    - Les projets de la source portent un trigramme (BDB, CIR…) qui n'est PAS
      stocké en base ; on garde en mémoire `trigramme → project.id` pour lier
      les tâches.
    - Les responsables de tâches sont rapprochés d'un User par nom normalisé.

Valeurs par défaut pour les trous de la source :
    - Projet sans date : [2026-05-01, 2028-12-31] (ajusté si une fin ou un
      jalon de fin est renseigné).
    - Tâche sans date : alignée sur la fenêtre du projet parent.
    - Jalon sans date : 2026-12-31.

Modèle des jalons : la source décrit des jalons « transverses » sans projet.
Depuis la migration 0008, un jalon doit être rattaché à au moins un projet
(INV-6). On crée donc un projet porteur unique « Jalons transverses » sous un
epic TVS, auquel tous ces jalons sont rattachés.
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
import unicodedata
from collections.abc import Iterable

import requests
from openpyxl import load_workbook

DEFAULT_PROJECT_START = dt.date(2026, 5, 1)
DEFAULT_PROJECT_END = dt.date(2028, 12, 31)
DEFAULT_MILESTONE_DATE = dt.date(2026, 12, 31)
DEFAULT_TASK_DURATION_DAYS = 30
EMAIL_DOMAIN = "lesfontaines.fr"  # TLD réel : email-validator refuserait .local


def norm(s: str | None) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return s.lower()


def slug_email(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", norm(name)) + "@" + EMAIL_DOMAIN


def to_date(v) -> dt.date | None:
    if v in (None, ""):
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, (int, float)):
        # Une année seule (ex : 2028.0) → 31 décembre de cette année.
        year = int(v)
        if 2000 <= year <= 2100:
            return dt.date(year, 12, 31)
    if isinstance(v, str):
        try:
            return dt.datetime.fromisoformat(v).date()
        except ValueError:
            pass
    return None


class Api:
    def __init__(self, base: str, token: str | None = None):
        self.base = base.rstrip("/")
        self.s = requests.Session()
        if token:
            self.s.headers["Authorization"] = f"Bearer {token}"

    def get(self, path: str, **params):
        r = self.s.get(self.base + path, params=params, timeout=10)
        r.raise_for_status()
        return r.json()

    def post(self, path: str, json):
        r = self.s.post(self.base + path, json=json, timeout=10)
        if r.status_code >= 400:
            return r.status_code, (r.json() if r.text else None)
        return r.status_code, r.json()

    def put(self, path: str, json):
        r = self.s.put(self.base + path, json=json, timeout=10)
        if r.status_code >= 400:
            return r.status_code, (r.json() if r.text else None)
        return r.status_code, r.json()


def login(api: Api, email: str, password: str) -> None:
    """Récupère un jeton via /api/auth/login et l'attache à la session."""
    r = api.s.post(
        api.base + "/api/auth/login",
        json={"email": email, "password": password},
        timeout=10,
    )
    if r.status_code != 200:
        raise SystemExit(f"Échec de connexion ({r.status_code}) : {r.text}")
    api.s.headers["Authorization"] = f"Bearer {r.json()['access_token']}"


def ensure_users(api: Api, names: Iterable[str]) -> dict[str, int]:
    """Crée les users manquants à partir des noms du classeur.

    Retourne `norm(nom) → user.id`.
    """
    existing_users = api.get("/api/users")
    out: dict[str, int] = {norm(u["nom"]): u["id"] for u in existing_users}
    created = 0
    for raw in names:
        nm = (raw or "").strip()
        if not nm or nm == "?":
            continue
        key = norm(nm)
        if key in out:
            continue
        payload = {
            "nom": nm,
            "email": slug_email(nm),
            "password": "changeme_dev",
            "role": "membre",
            "actif": True,
        }
        code, body = api.post("/api/users", payload)
        if code >= 400:
            print(f"  ! user {nm!r} refusé : {body}")
            continue
        out[key] = body["id"]
        created += 1
    print(f"users : {created} créés (total {len(out)})")
    return out


def ensure_epic(
    api: Api, trigramme: str, nom: str, *, statut: str = "idee", critere: str | None = None
) -> None:
    existing = api.get("/api/epics")
    if any(e["trigramme"] == trigramme for e in existing):
        return
    payload = {
        "trigramme": trigramme,
        "nom": nom,
        "critere_reussite": critere,
        "statut": statut,
        "categorie": "operationnel",
    }
    code, body = api.post("/api/epics", payload)
    if code >= 400:
        print(f"  ! epic {trigramme} refusé : {body}")


def import_projects(api: Api, wb) -> dict[str, int]:
    """Importe « Projets » + « Projet non plannifiés ». Retourne `trig → id`."""
    ws = wb["Projets"]
    rows = list(ws.iter_rows(values_only=True))[1:]

    existing = api.get("/api/projects")
    out: dict[str, int] = {}
    by_name: dict[tuple[str, str], int] = {
        (p["epic_trigramme"], p["nom"]): p["id"] for p in existing
    }

    created = skipped = refused = 0
    for r in rows:
        if not any(c not in (None, "") for c in r):
            continue
        nom = (r[0] or "").strip()
        fin_prevu = to_date(r[1])
        jalon_max = to_date(r[2])
        raison = (r[3] or "").strip() or None
        trig = (r[4] or "").strip()
        epic = (r[5] or "").strip()
        rappel = (r[6] or "").strip() if len(r) > 6 and r[6] else ""
        termine = bool(r[7]) if len(r) > 7 else False

        if not nom or not epic:
            continue

        # « Rappel du titre de l'épic » donne le nom complet de l'epic ; on
        # s'en sert pour les epics créés à la volée (sinon le nom = trigramme).
        ensure_epic(api, epic.upper(), rappel or epic.upper())

        date_fin = fin_prevu or jalon_max or DEFAULT_PROJECT_END
        if date_fin < DEFAULT_PROJECT_START:
            date_fin = DEFAULT_PROJECT_END

        key = (epic.upper(), nom)
        if key in by_name:
            if trig:
                out[trig.upper()] = by_name[key]
            skipped += 1
            continue

        desc_parts = []
        if rappel:
            desc_parts.append(rappel)
        if raison:
            desc_parts.append(f"Raison fin : {raison}")
        description = " · ".join(desc_parts) or None

        payload = {
            "epic_trigramme": epic.upper(),
            "nom": nom,
            "description": description,
            "date_debut": DEFAULT_PROJECT_START.isoformat(),
            "date_fin": date_fin.isoformat(),
            "statut": "realise" if termine else "en_cours",
        }
        code, body = api.post("/api/projects", payload)
        if code >= 400:
            refused += 1
            print(f"  ! projet {nom!r} (epic {epic}) refusé : {body}")
            continue
        if trig:
            out[trig.upper()] = body["id"]
        by_name[key] = body["id"]
        created += 1

    # Projets non planifiés → epic NPL. Cet onglet n'a pas toujours de ligne
    # d'en-tête (le vrai classeur commence directement par une donnée) : on ne
    # saute donc pas la première ligne, on écarte seulement les intitulés de
    # colonne s'ils sont présents.
    ensure_epic(api, "NPL", "Projets non planifiés (à classer)", statut="idee")
    entetes = {"nom", "nom du projet", "projet", "projets"}
    np_created = 0
    for r in wb["Projet non plannifiés"].iter_rows(values_only=True):
        nom = (r[0] or "").strip() if r and r[0] else ""
        if not nom or norm(nom) in entetes or ("NPL", nom) in by_name:
            continue
        payload = {
            "epic_trigramme": "NPL",
            "nom": nom,
            "date_debut": DEFAULT_PROJECT_START.isoformat(),
            "date_fin": DEFAULT_PROJECT_END.isoformat(),
            "statut": "prevu",
        }
        code, body = api.post("/api/projects", payload)
        if code >= 400:
            print(f"  ! projet non planifié {nom!r} refusé : {body}")
            continue
        by_name[("NPL", nom)] = body["id"]
        np_created += 1

    print(
        f"projets : {created} créés, {skipped} déjà présents, "
        f"{refused} refusés, {np_created} non planifiés"
    )
    return out


def import_tasks(
    api: Api, wb, projects_by_trig: dict[str, int], users: dict[str, int]
) -> None:
    ws = wb["Detail des tache de projet"]
    rows = list(ws.iter_rows(values_only=True))[1:]

    projects = api.get("/api/projects")
    win = {
        p["id"]: (dt.date.fromisoformat(p["date_debut"]), dt.date.fromisoformat(p["date_fin"]))
        for p in projects
    }

    existing = api.get("/api/tasks")
    by_key: set[tuple[int, str]] = {(t["projet_id"], t["nom"]) for t in existing}

    created = refused = skipped = no_project = 0
    for r in rows:
        if not any(c not in (None, "") for c in r):
            continue
        trig = (r[0] or "").strip().upper()
        nom = (r[2] or "").strip() if len(r) > 2 else ""
        date_debut = to_date(r[3]) if len(r) > 3 else None
        date_fin_src = to_date(r[4]) if len(r) > 4 else None
        responsable = (r[6] or "").strip() if len(r) > 6 and r[6] else ""
        termine = bool(r[9]) if len(r) > 9 else False

        if not nom or not trig:
            continue

        pid = projects_by_trig.get(trig)
        if pid is None:
            no_project += 1
            continue
        if (pid, nom) in by_key:
            skipped += 1
            continue

        pstart, pend = win[pid]
        ts = date_debut or pstart
        te = date_fin_src or (ts + dt.timedelta(days=DEFAULT_TASK_DURATION_DAYS))
        # On aligne les dates sur la fenêtre du projet pour un rendu propre
        # (INV-9 est retiré : ce clip est du confort, pas une contrainte API).
        ts = max(pstart, min(pend, ts))
        te = max(ts, min(pend, te))

        payload = {
            "projet_id": pid,
            "nom": nom,
            "date_debut": ts.isoformat(),
            "date_fin": te.isoformat(),
            "statut": "archive" if termine else "ouvert",
            "responsable_id": users.get(norm(responsable)),
        }
        code, body = api.post("/api/tasks", payload)
        if code >= 400:
            refused += 1
            print(f"  ! tâche {nom!r} (projet {trig}) refusée : {body}")
            continue
        by_key.add((pid, nom))
        created += 1

    print(
        f"tâches : {created} créées, {skipped} déjà présentes, "
        f"{refused} refusées, {no_project} sans projet identifié"
    )


def ensure_transverse_project(api: Api) -> int | None:
    """Projet porteur unique des jalons transverses (voir docstring, INV-6).

    Idempotent : réutilise le projet s'il existe déjà.
    """
    ensure_epic(
        api, "TVS", "Jalons transverses (suivi général)",
        statut="actif", critere="Suivi des jalons réglementaires et événementiels",
    )
    nom = "Jalons transverses"
    for p in api.get("/api/projects"):
        if p["epic_trigramme"] == "TVS" and p["nom"] == nom:
            return p["id"]
    code, body = api.post("/api/projects", {
        "epic_trigramme": "TVS",
        "nom": nom,
        "description": "Projet porteur des jalons sans rattachement propre.",
        "date_debut": DEFAULT_PROJECT_START.isoformat(),
        "date_fin": DEFAULT_PROJECT_END.isoformat(),
        "statut": "en_cours",
    })
    if code >= 400:
        print(f"  ! projet porteur des jalons refusé : {body}")
        return None
    return body["id"]


def import_milestones(api: Api, wb) -> None:
    if "Jalons" not in wb.sheetnames:
        print("jalons : pas d'onglet « Jalons » dans la source — ignoré")
        return
    porteur_id = ensure_transverse_project(api)
    if porteur_id is None:
        print("jalons : projet porteur indisponible — import ignoré")
        return

    existing = {m["nom"] for m in api.get("/api/milestones")}
    created = refused = 0
    for r in list(wb["Jalons"].iter_rows(values_only=True))[1:]:  # saute l'en-tête
        if not r or not r[0]:
            continue
        nom = (r[0] or "").strip()
        if not nom or nom in existing:
            continue
        date = (to_date(r[1]) if len(r) > 1 else None) or DEFAULT_MILESTONE_DATE
        code, body = api.post("/api/milestones", {
            "nom": nom,
            "date": date.isoformat(),
            "atteint": False,
            "project_ids": [porteur_id],
        })
        if code >= 400:
            refused += 1
            print(f"  ! jalon {nom!r} refusé : {body}")
            continue
        existing.add(nom)
        created += 1
    print(f"jalons : {created} créés, {refused} refusés")


def main() -> int:
    ap = argparse.ArgumentParser(description="Import du classeur source dans l'API.")
    ap.add_argument("--api", default="http://localhost:8080")
    ap.add_argument("--xlsx", default="data/source.xlsx")
    ap.add_argument("--token", help="Jeton JWT (sinon --email/--password, sinon AUTH_DISABLED)")
    ap.add_argument("--email", help="Email admin pour se connecter et obtenir un jeton")
    ap.add_argument("--password", help="Mot de passe associé à --email")
    args = ap.parse_args()

    print(f"API    : {args.api}")
    print(f"Source : {args.xlsx}\n")

    api = Api(args.api, token=args.token)
    if not args.token and args.email and args.password:
        login(api, args.email, args.password)

    wb = load_workbook(args.xlsx, data_only=True)

    # 1. Users : chargés de projets + responsables uniques des tâches
    names: set[str] = set()
    for r in list(wb["Chargés de projets"].iter_rows(values_only=True))[1:]:
        if r and r[0]:
            names.add(r[0])
    for r in list(wb["Detail des tache de projet"].iter_rows(values_only=True))[1:]:
        if len(r) > 6 and r[6]:
            names.add(r[6])
    users = ensure_users(api, names)

    # 2. Projets (crée les epics manquants à la volée)
    projects_by_trig = import_projects(api, wb)

    # 3. Tâches
    import_tasks(api, wb, projects_by_trig, users)

    # 4. Jalons
    import_milestones(api, wb)

    # 5. Résumé
    print("\n=== Résumé en base ===")
    for path in ("/api/epics", "/api/projects", "/api/tasks", "/api/milestones", "/api/users"):
        print(f"  {path} : {len(api.get(path))}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

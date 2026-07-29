"""Export du planning au format du classeur source — la réciproque de l'import.

Sert à emporter l'état courant vers une autre installation (typiquement local →
production) sans passer par un dump PostgreSQL, et à réalimenter le tableur de
suivi après des corrections faites dans l'application.

Le format visé est **exactement** celui que lit `import_xlsx.py` : sa docstring
fait foi, et `test_export_xlsx.py` éprouve le circuit complet
(export → import → comparaison) plutôt que de faire confiance à cette phrase.

CE QUE LE FORMAT NE SAIT PAS PORTER
-----------------------------------
Un classeur source n'est pas un dump : il décrit ce que la pisciculture saisit,
pas tout ce que l'application stocke. Ne survivent donc pas à un aller-retour :

- les **dépendances**, **équipes**, **allocations** et **mesures** — le format
  n'a pas d'onglet pour elles ;
- les **métadonnées d'epic** (critère de réussite, catégorie, couleur, échéance) :
  seul le nom transite, via la colonne « Rappel ». En production, ce sont le CSV
  de seed et la saisie qui les portent ;
- le **rattachement réel des jalons** à leurs projets : l'import les regroupe
  tous sous le projet porteur « Jalons transverses » (INV-6 exige au moins un
  projet) ;
- la **date de début des projets**, absente de la source : l'import la fixe à
  `DEFAULT_PROJECT_START`.

`main()` annonce ce qui est perdu **pour les données réellement exportées**, et
se tait sur le reste : un avertissement qui se déclenche toujours n'est plus lu.
Pour une copie fidèle, c'est `docs/RESTORE.md` (dump PostgreSQL) qu'il faut.
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.dependency import Dependency
from app.models.epic import Epic
from app.models.equipe import Equipe
from app.models.measure import Measure
from app.models.milestone import Milestone
from app.models.project import Project, ProjectStatus
from app.models.task import Task, TaskStatus
from app.models.user import User

#: Séparateur posé par l'import entre le rappel du titre d'epic et la raison de
#: la date de fin, quand il fabrique `Project.description`. On le défait ici pour
#: rendre à chaque colonne son contenu.
_SEP_RAISON = " · Raison fin : "

#: Epic des projets non planifiés : l'import les lit dans un onglet à part et
#: leur donne le statut `prevu`. Les réémettre dans « Projets » les ferait
#: revenir en `en_cours` — le statut se perdrait sans que rien ne le signale.
_EPIC_NON_PLANIFIES = "NPL"

#: En-têtes du classeur source, reproduits à l'identique. Ils ne sont pas relus
#: par l'import (qui saute la première ligne et travaille par indice), mais un
#: humain ouvre ce fichier : des colonnes anonymes le rendraient illisible.
_ENTETES_PROJETS = [
    "Nom", "fin prévu", "Jalon de fin maximum", "Raison de la date de fin",
    "Trigramme", "Epic lié", "Rappel du titre de l'épic", "Terminé",
]
_ENTETES_TACHES = [
    "Projet lié", "Rappel", "Nom de la tache", "Date de début", "Jalon maximum",
    "Durée", "Responsable", "Equipe", "Materiel", "Terminé",
]


@dataclass
class RapportExport:
    utilisateurs: int = 0
    projets: int = 0
    projets_non_planifies: int = 0
    taches: int = 0
    jalons: int = 0
    #: Ce qui n'a pas pu être écrit, avec son motif. Vide = tout est passé.
    omissions: list[str] = field(default_factory=list)


def _trigramme(numero: int) -> str:
    """Identifiant de projet interne au fichier, seul lien projet ↔ tâches.

    Les trigrammes du classeur d'origine (BDB, CIR…) ne sont **pas** stockés en
    base : ils ne servaient qu'à rattacher les tâches pendant l'import, et rien
    ne les a conservés. On en fabrique donc de nouveaux.

    Sans conséquence à la réimportation : l'import reconnaît un projet existant
    par `(epic, nom)`, jamais par son trigramme. Deux exports successifs restent
    donc idempotents même si les trigrammes changent.
    """
    return f"P{numero:03d}"


def _rappel_et_raison(description: str | None, epic_nom: str) -> tuple[str, str | None]:
    """Défait `description` en (rappel du titre d'epic, raison de la date de fin).

    Le rappel émis est le nom RÉEL de l'epic, pas le préfixe trouvé dans la
    description : c'est le sens de cette colonne dans la source, et un epic
    renommé depuis l'import doit ressortir sous son nom actuel.
    """
    if description and _SEP_RAISON in description:
        return epic_nom, description.split(_SEP_RAISON, 1)[1].strip() or None
    return epic_nom, None


def exporter_classeur(db: Session) -> tuple[bytes, RapportExport]:
    rapport = RapportExport()
    wb = Workbook()

    # --- Chargés de projets ---------------------------------------------
    # Seulement les comptes ACTIFS : réémettre un compte désactivé le ferait
    # revenir actif à l'import (`"actif": True` en dur), ressuscitant en silence
    # un accès qu'on avait retiré.
    ws = wb.active
    ws.title = "Chargés de projets"
    ws.append(["Nom"])
    for (nom,) in db.execute(select(User.nom).where(User.actif).order_by(User.nom)):
        ws.append([nom])
        rapport.utilisateurs += 1

    epics = {e.trigramme: e.nom for e in db.execute(select(Epic)).scalars()}
    projets = list(db.execute(select(Project).order_by(Project.id)).scalars())

    # --- Projets ---------------------------------------------------------
    trig_par_projet: dict[int, str] = {}
    ws = wb.create_sheet("Projets")
    ws.append(_ENTETES_PROJETS)
    for numero, p in enumerate(
        (p for p in projets if p.epic_trigramme != _EPIC_NON_PLANIFIES), start=1
    ):
        trig = _trigramme(numero)
        trig_par_projet[p.id] = trig
        rappel, raison = _rappel_et_raison(p.description, epics.get(p.epic_trigramme, ""))
        ws.append([
            p.nom,
            p.date_fin,
            None,  # jalon de fin max : l'import retombe sur « fin prévu »
            raison,
            trig,
            p.epic_trigramme,
            rappel,
            True if p.statut == ProjectStatus.realise else None,
        ])
        rapport.projets += 1

    # --- Projets non planifiés -------------------------------------------
    ws = wb.create_sheet("Projet non plannifiés")  # coquille du classeur d'origine
    ws.append(["Nom"])
    for p in projets:
        if p.epic_trigramme == _EPIC_NON_PLANIFIES:
            ws.append([p.nom])
            rapport.projets_non_planifies += 1

    # --- Tâches -----------------------------------------------------------
    ws = wb.create_sheet("Detail des tache de projet")
    ws.append(_ENTETES_TACHES)
    responsables = {u.id: u.nom for u in db.execute(select(User)).scalars()}
    for t in db.execute(select(Task).order_by(Task.id)).scalars():
        trig = trig_par_projet.get(t.projet_id)
        if trig is None:
            # Le projet porteur n'a pas de ligne dans « Projets » (non planifié),
            # donc rien à quoi rattacher la tâche. Le dire : une tâche disparue
            # sans un mot est exactement ce qu'un export ne doit jamais faire.
            rapport.omissions.append(
                f"tâche {t.nom!r} : son projet est non planifié, "
                "et cet onglet ne porte pas de trigramme"
            )
            continue
        ws.append([
            trig,
            None,
            t.nom,
            t.date_debut,
            t.date_fin,
            None,  # durée : l'import n'en tient pas compte, les dates priment
            responsables.get(t.responsable_id) if t.responsable_id else None,
            None,
            None,
            True if t.statut == TaskStatus.archive else None,
        ])
        rapport.taches += 1

    # --- Jalons ------------------------------------------------------------
    jalons = list(db.execute(select(Milestone).order_by(Milestone.date)).scalars())
    if jalons:
        ws = wb.create_sheet("Jalons")
        ws.append(["Nom", "Date"])
        for m in jalons:
            ws.append([m.nom, m.date])
            rapport.jalons += 1
        rapport.omissions.append(
            f"{len(jalons)} jalon(s) : leur rattachement aux projets n'est pas "
            "représentable — l'import les regroupera sous « Jalons transverses »"
        )

    # Ce que le format ignore complètement. On ne le signale que si la base en
    # contient : un avertissement qui se déclenche toujours n'est plus lu.
    for modele, libelle in ((Dependency, "dépendances"), (Equipe, "équipes"), (Measure, "mesures")):
        nb = db.execute(select(func.count()).select_from(modele)).scalar_one()
        if nb:
            rapport.omissions.append(
                f"{nb} {libelle} : le format source n'a pas d'onglet pour elles"
            )

    tampon = BytesIO()
    wb.save(tampon)
    return tampon.getvalue(), rapport


def main() -> None:
    from app.database import SessionLocal

    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out", required=True, help="chemin du classeur à écrire")
    args = ap.parse_args()

    with SessionLocal() as db:
        contenu, rapport = exporter_classeur(db)
    Path(args.out).write_bytes(contenu)

    print(f"Classeur écrit : {args.out} ({len(contenu)} octets)")
    print(f"  chargés de projets  : {rapport.utilisateurs}")
    print(f"  projets             : {rapport.projets}")
    print(f"  projets non planifiés: {rapport.projets_non_planifies}")
    print(f"  tâches              : {rapport.taches}")
    print(f"  jalons              : {rapport.jalons}")
    if rapport.omissions:
        print("\nNON REPRÉSENTABLE dans ce format :")
        for m in rapport.omissions:
            print(f"  - {m}")


if __name__ == "__main__":
    main()

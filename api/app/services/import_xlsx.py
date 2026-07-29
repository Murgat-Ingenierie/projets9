"""Import du classeur source — logique métier, sans transport.

Déplacée telle quelle depuis `scripts/import_data.py` : elle n'a jamais parlé
qu'à une interface minuscule (`get`/`post`), ce qui la rend indifférente au
canal. Deux implémentations la nourrissent aujourd'hui :

- `ClientEnProcess` (app/services/import_client.py) — appels directs aux
  fonctions de route, utilisé par l'endpoint d'import ;
- l'ancien client HTTP du script, tant qu'il existe.

**Ce qui compte** : les écritures passent toujours par les routes, donc par les
invariants. Une ligne invalide est refusée et rapportée, jamais insérée en
douce — c'est ce qui a fait remonter les 9 refus INV-10 du premier import réel.

Format attendu du classeur : voir la docstring de `scripts/import_data.py`.
"""

from __future__ import annotations

import datetime as dt
import re
import unicodedata
from collections.abc import Iterable
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any, Protocol


class ClientImport(Protocol):
    """Le contrat minimal dont l'import a besoin — et rien de plus.

    C'est ce qui rend la logique indifférente au transport : `ClientEnProcess`
    (appels directs aux routes) et l'ancien client HTTP l'honorent tous deux.
    Les écritures renvoient `(code, corps)` plutôt que de lever : un refus
    d'invariant est une information à rapporter, pas une panne.
    """

    def get(self, chemin: str, **params: Any) -> Any: ...

    def post(self, chemin: str, json: dict) -> tuple[int, Any]: ...


#: Journal de l'import en cours. Les fonctions déplacées depuis le script
#: écrivaient leur bilan sur la sortie standard ; une interface a besoin de le
#: RÉCUPÉRER. On conserve leurs signatures (donc leur logique intacte) en
#: passant par ce journal, remis à zéro à chaque import.
_JOURNAL: dict[str, Any] = {}


def _motif(body: Any) -> str:
    """Message lisible d'un refus : le code d'invariant si présent."""
    if isinstance(body, dict):
        d = body.get("detail", body)
        if isinstance(d, dict):
            return f"{d.get('code', '?')} — {d.get('message', '')}"
        return str(d)
    return str(body)

#: Valeurs de repli quand la source ne porte pas l'information. Le classeur ne
#: renseigne pas de date de début de projet : sans ces défauts, rien ne passerait
#: la validation. C'est aussi pourquoi le planning importé paraît « plat ».
DEFAULT_PROJECT_START = dt.date(2026, 5, 1)
DEFAULT_PROJECT_END = dt.date(2028, 12, 31)
DEFAULT_MILESTONE_DATE = dt.date(2026, 12, 31)
DEFAULT_TASK_DURATION_DAYS = 30
EMAIL_DOMAIN = "lesfontaines.fr"  # TLD réel : email-validator refuserait .local


def norm(s: str | None) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return s.lower()


def slug_email(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", norm(name)) + "@" + EMAIL_DOMAIN


def to_date(v) -> dt.date | None:
    if v in (None, ""):
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, (int, float)):
        # Une année seule (ex : 2028.0) → 31 décembre de cette année.
        year = int(v)
        if 2000 <= year <= 2100:
            return dt.date(year, 12, 31)
    if isinstance(v, str):
        try:
            return dt.datetime.fromisoformat(v).date()
        except ValueError:
            pass
    return None


def ensure_users(api: ClientImport, names: Iterable[str]) -> dict[str, int]:
    """Crée les users manquants à partir des noms du classeur.

    Retourne `norm(nom) → user.id`.
    """
    existing_users = api.get("/api/users")
    out: dict[str, int] = {norm(u["nom"]): u["id"] for u in existing_users}
    created = 0
    for raw in names:
        nm = (raw or "").strip()
        if not nm or nm == "?":
            continue
        key = norm(nm)
        if key in out:
            continue
        payload = {
            "nom": nm,
            "email": slug_email(nm),
            "role": "membre",
            "actif": True,
        }
        code, body = api.post("/api/users", payload)
        if code >= 400:
            _JOURNAL["refus"].append(f"utilisateur {nm!r} : {_motif(body)}")
            continue
        out[key] = body["id"]
        created += 1
    _JOURNAL["utilisateurs_crees"] = created
    return out


def ensure_epic(
    api: ClientImport, trigramme: str, nom: str, *, statut: str = "idee", critere: str | None = None
) -> None:
    existing = api.get("/api/epics")
    if any(e["trigramme"] == trigramme for e in existing):
        return
    payload = {
        "trigramme": trigramme,
        "nom": nom,
        "critere_reussite": critere,
        "statut": statut,
        "categorie": "operationnel",
    }
    code, body = api.post("/api/epics", payload)
    if code >= 400:
        _JOURNAL["refus"].append(f"epic {trigramme} : {_motif(body)}")


def import_projects(api: ClientImport, wb) -> dict[str, int]:
    """Importe « Projets » + « Projet non plannifiés ». Retourne `trig → id`."""
    ws = wb["Projets"]
    rows = list(ws.iter_rows(values_only=True))[1:]

    existing = api.get("/api/projects")
    out: dict[str, int] = {}
    by_name: dict[tuple[str, str], int] = {
        (p["epic_trigramme"], p["nom"]): p["id"] for p in existing
    }

    created = skipped = refused = 0
    for r in rows:
        if not any(c not in (None, "") for c in r):
            continue
        nom = (r[0] or "").strip()
        fin_prevu = to_date(r[1])
        jalon_max = to_date(r[2])
        raison = (r[3] or "").strip() or None
        trig = (r[4] or "").strip()
        epic = (r[5] or "").strip()
        rappel = (r[6] or "").strip() if len(r) > 6 and r[6] else ""
        termine = bool(r[7]) if len(r) > 7 else False

        if not nom or not epic:
            continue

        # « Rappel du titre de l'épic » donne le nom complet de l'epic ; on
        # s'en sert pour les epics créés à la volée (sinon le nom = trigramme).
        ensure_epic(api, epic.upper(), rappel or epic.upper())

        date_fin = fin_prevu or jalon_max or DEFAULT_PROJECT_END
        if date_fin < DEFAULT_PROJECT_START:
            date_fin = DEFAULT_PROJECT_END

        key = (epic.upper(), nom)
        if key in by_name:
            if trig:
                out[trig.upper()] = by_name[key]
            skipped += 1
            continue

        desc_parts = []
        if rappel:
            desc_parts.append(rappel)
        if raison:
            desc_parts.append(f"Raison fin : {raison}")
        description = " · ".join(desc_parts) or None

        payload = {
            "epic_trigramme": epic.upper(),
            "nom": nom,
            "description": description,
            "date_debut": DEFAULT_PROJECT_START.isoformat(),
            "date_fin": date_fin.isoformat(),
            "statut": "realise" if termine else "en_cours",
        }
        code, body = api.post("/api/projects", payload)
        if code >= 400:
            refused += 1
            _JOURNAL["refus"].append(f"projet {nom!r} (epic {epic}) : {_motif(body)}")
            continue
        if trig:
            out[trig.upper()] = body["id"]
        by_name[key] = body["id"]
        created += 1

    # Projets non planifiés → epic NPL. Cet onglet n'a pas toujours de ligne
    # d'en-tête (le vrai classeur commence directement par une donnée) : on ne
    # saute donc pas la première ligne, on écarte seulement les intitulés de
    # colonne s'ils sont présents.
    # Onglet OPTIONNEL : le classeur réel le porte, un export partiel non. Depuis
    # que l'import est exposé dans l'interface, il reçoit des fichiers
    # quelconques — une absence doit être ignorée, pas remonter en KeyError.
    # (Le nom contient la coquille du classeur d'origine : « plannifiés ».)
    ONGLET_NON_PLANIFIES = "Projet non plannifiés"
    np_created = 0
    if ONGLET_NON_PLANIFIES not in wb.sheetnames:
        _JOURNAL.update(projets_crees=created, projets_deja=skipped,
                        projets_refuses=refused, projets_non_planifies=0)
        return out

    ensure_epic(api, "NPL", "Projets non planifiés (à classer)", statut="idee")
    entetes = {"nom", "nom du projet", "projet", "projets"}
    for r in wb[ONGLET_NON_PLANIFIES].iter_rows(values_only=True):
        nom = (r[0] or "").strip() if r and r[0] else ""
        if not nom or norm(nom) in entetes or ("NPL", nom) in by_name:
            continue
        payload = {
            "epic_trigramme": "NPL",
            "nom": nom,
            "date_debut": DEFAULT_PROJECT_START.isoformat(),
            "date_fin": DEFAULT_PROJECT_END.isoformat(),
            "statut": "prevu",
        }
        code, body = api.post("/api/projects", payload)
        if code >= 400:
            _JOURNAL["refus"].append(f"projet non planifié {nom!r} : {_motif(body)}")
            continue
        by_name[("NPL", nom)] = body["id"]
        np_created += 1

    _JOURNAL.update(projets_crees=created, projets_deja=skipped,
                    projets_refuses=refused, projets_non_planifies=np_created)
    return out


def import_tasks(
    api: ClientImport, wb, projects_by_trig: dict[str, int], users: dict[str, int]
) -> None:
    ws = wb["Detail des tache de projet"]
    rows = list(ws.iter_rows(values_only=True))[1:]

    projects = api.get("/api/projects")
    win = {
        p["id"]: (dt.date.fromisoformat(p["date_debut"]), dt.date.fromisoformat(p["date_fin"]))
        for p in projects
    }

    existing = api.get("/api/tasks")
    by_key: set[tuple[int, str]] = {(t["projet_id"], t["nom"]) for t in existing}

    created = refused = skipped = no_project = 0
    for r in rows:
        if not any(c not in (None, "") for c in r):
            continue
        trig = (r[0] or "").strip().upper()
        nom = (r[2] or "").strip() if len(r) > 2 else ""
        date_debut = to_date(r[3]) if len(r) > 3 else None
        date_fin_src = to_date(r[4]) if len(r) > 4 else None
        responsable = (r[6] or "").strip() if len(r) > 6 and r[6] else ""
        termine = bool(r[9]) if len(r) > 9 else False

        if not nom or not trig:
            continue

        pid = projects_by_trig.get(trig)
        if pid is None:
            no_project += 1
            continue
        if (pid, nom) in by_key:
            skipped += 1
            continue

        pstart, pend = win[pid]
        ts = date_debut or pstart
        te = date_fin_src or (ts + dt.timedelta(days=DEFAULT_TASK_DURATION_DAYS))
        # On aligne les dates sur la fenêtre du projet pour un rendu propre
        # (INV-9 est retiré : ce clip est du confort, pas une contrainte API).
        ts = max(pstart, min(pend, ts))
        te = max(ts, min(pend, te))

        payload = {
            "projet_id": pid,
            "nom": nom,
            "date_debut": ts.isoformat(),
            "date_fin": te.isoformat(),
            "statut": "archive" if termine else "ouvert",
            "responsable_id": users.get(norm(responsable)),
        }
        code, body = api.post("/api/tasks", payload)
        if code >= 400:
            refused += 1
            _JOURNAL["refus"].append(f"tâche {nom!r} (projet {trig}) : {_motif(body)}")
            continue
        by_key.add((pid, nom))
        created += 1

    _JOURNAL.update(taches_creees=created, taches_deja=skipped,
                    taches_refusees=refused, taches_sans_projet=no_project)


def ensure_transverse_project(api: ClientImport) -> int | None:
    """Projet porteur unique des jalons transverses (voir docstring, INV-6).

    Idempotent : réutilise le projet s'il existe déjà.
    """
    ensure_epic(
        api, "TVS", "Jalons transverses (suivi général)",
        statut="actif", critere="Suivi des jalons réglementaires et événementiels",
    )
    nom = "Jalons transverses"
    for p in api.get("/api/projects"):
        if p["epic_trigramme"] == "TVS" and p["nom"] == nom:
            return p["id"]
    code, body = api.post("/api/projects", {
        "epic_trigramme": "TVS",
        "nom": nom,
        "description": "Projet porteur des jalons sans rattachement propre.",
        "date_debut": DEFAULT_PROJECT_START.isoformat(),
        "date_fin": DEFAULT_PROJECT_END.isoformat(),
        "statut": "en_cours",
    })
    if code >= 400:
        _JOURNAL["refus"].append(f"projet porteur des jalons : {_motif(body)}")
        return None
    return body["id"]


def import_milestones(api: ClientImport, wb) -> None:
    if "Jalons" not in wb.sheetnames:
        _JOURNAL["jalons"] = "pas d'onglet « Jalons » dans la source — ignoré"
        return
    porteur_id = ensure_transverse_project(api)
    if porteur_id is None:
        _JOURNAL["jalons"] = "projet porteur indisponible — import ignoré"
        return

    existing = {m["nom"] for m in api.get("/api/milestones")}
    created = refused = 0
    for r in list(wb["Jalons"].iter_rows(values_only=True))[1:]:  # saute l'en-tête
        if not r or not r[0]:
            continue
        nom = (r[0] or "").strip()
        if not nom or nom in existing:
            continue
        date = (to_date(r[1]) if len(r) > 1 else None) or DEFAULT_MILESTONE_DATE
        code, body = api.post("/api/milestones", {
            "nom": nom,
            "date": date.isoformat(),
            "atteint": False,
            "project_ids": [porteur_id],
        })
        if code >= 400:
            refused += 1
            _JOURNAL["refus"].append(f"jalon {nom!r} : {_motif(body)}")
            continue
        existing.add(nom)
        created += 1
    _JOURNAL["jalons"] = f"{created} créés, {refused} refusés"




# --- orchestration ------------------------------------------------------------


@dataclass
class RapportImport:
    """Ce que l'import a fait, et surtout ce qu'il a REFUSÉ.

    Les refus ne sont pas des erreurs de l'import : ce sont des invariants qui
    ont fait leur travail. Ils doivent donc remonter à l'utilisateur, avec leur
    motif — c'est ainsi qu'on a découvert que 9 projets visaient une échéance
    postérieure à celle de leur epic.
    """

    utilisateurs_crees: int = 0
    projets_crees: int = 0
    projets_deja_presents: int = 0
    refus: list[str] = field(default_factory=list)
    projets_non_planifies: int = 0
    taches_creees: int = 0
    taches_deja_presentes: int = 0
    taches_sans_projet: int = 0
    jalons: str = ""
    totaux: dict[str, int] = field(default_factory=dict)


def importer_classeur(contenu: bytes, client) -> RapportImport:
    """Importe un classeur (contenu binaire) via `client` (get/post).

    `client` est soit `ClientEnProcess`, soit l'ancien client HTTP : la logique
    est identique, seul le transport change.
    """
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(contenu), data_only=True)
    _JOURNAL.clear()
    _JOURNAL["refus"] = []
    _JOURNAL["jalons"] = ""
    rapport = RapportImport()

    manquants = [f for f in ("Chargés de projets", "Projets", "Detail des tache de projet")
                 if f not in wb.sheetnames]
    if manquants:
        # Message actionnable : l'utilisateur a probablement déposé le mauvais
        # fichier, ou un export partiel.
        raise ValueError(
            "Onglet(s) manquant(s) dans le classeur : " + ", ".join(manquants)
            + ". Onglets trouvés : " + ", ".join(wb.sheetnames)
        )

    noms: set[str] = set()
    for r in list(wb["Chargés de projets"].iter_rows(values_only=True))[1:]:
        if r and r[0]:
            noms.add(r[0])
    for r in list(wb["Detail des tache de projet"].iter_rows(values_only=True))[1:]:
        if len(r) > 6 and r[6]:
            noms.add(r[6])

    utilisateurs = ensure_users(client, noms)
    projets_par_trigramme = import_projects(client, wb)
    import_tasks(client, wb, projets_par_trigramme, utilisateurs)
    import_milestones(client, wb)

    rapport.utilisateurs_crees = _JOURNAL.get("utilisateurs_crees", 0)
    rapport.projets_crees = _JOURNAL.get("projets_crees", 0)
    rapport.projets_deja_presents = _JOURNAL.get("projets_deja", 0)
    rapport.projets_non_planifies = _JOURNAL.get("projets_non_planifies", 0)
    rapport.taches_creees = _JOURNAL.get("taches_creees", 0)
    rapport.taches_deja_presentes = _JOURNAL.get("taches_deja", 0)
    rapport.taches_sans_projet = _JOURNAL.get("taches_sans_projet", 0)
    rapport.jalons = _JOURNAL.get("jalons", "")
    # Un seul champ de refus, toutes natures confondues : ce qui compte pour
    # l'utilisateur est CE QUI N'EST PAS PASSÉ, pas la catégorie de la ligne.
    rapport.refus = list(_JOURNAL.get("refus", []))

    rapport.totaux = {
        chemin.rsplit("/", 1)[-1]: len(client.get(chemin))
        for chemin in ("/api/epics", "/api/projects", "/api/tasks", "/api/milestones", "/api/users")
    }
    return rapport

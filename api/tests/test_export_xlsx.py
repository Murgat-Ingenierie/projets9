"""Export du planning au format source — éprouvé par l'ALLER-RETOUR.

Un test qui vérifierait « le fichier contient les bons onglets » ne prouverait
rien : c'est l'import qui décide si un classeur est lisible, et il travaille par
INDICE de colonne. Une colonne insérée au bon endroit du mauvais onglet passerait
la relecture d'un humain comme celle d'un test structurel.

On exporte donc, puis on réimporte dans une base vierge par le chemin réel
(`ClientEnProcess`, donc les routes, donc les invariants), et on compare. Ce qui
diverge est soit un défaut, soit une perte connue du format — et dans ce dernier
cas, `RapportExport.omissions` doit l'avoir annoncée.
"""

from __future__ import annotations

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import sessionmaker

from app.models.project import Project
from app.models.task import Task
from app.models.user import User
from app.services.export_xlsx import exporter_classeur
from app.services.import_client import ClientEnProcess
from app.services.import_xlsx import importer_classeur


@pytest.fixture
def planning(client: TestClient, auth, fabrique):
    """Un planning représentatif : deux epics, des tâches datées, un responsable."""
    fabrique.epic("ABC")
    fabrique.epic("DEF", nom="Deuxième epic")
    p1 = fabrique.projet("ABC", nom="Bassin nord")
    p2 = fabrique.projet("DEF", nom="Filtration", date_fin="2026-09-30")
    fabrique.tache(p1["id"], nom="Poser les capteurs")
    fabrique.tache(p1["id"], nom="Calibrer", date_debut="2026-08-05", date_fin="2026-08-20")
    fabrique.tache(p2["id"], nom="Étude hydraulique")
    return {"p1": p1, "p2": p2}


def _relire(contenu: bytes, db):
    """Réimporte dans `db` par le chemin réel, et renvoie le rapport d'import."""
    admin = db.query(User).filter(User.role == "admin").first()
    return importer_classeur(contenu, ClientEnProcess(db, admin))


# --- l'aller-retour ----------------------------------------------------------


def test_aller_retour_preserve_projets_et_taches(
    client, auth, planning, session_factory: sessionmaker
):
    """Le cœur : ce qui sort doit rentrer, à l'identique."""
    source = session_factory()
    contenu, rapport = exporter_classeur(source)
    avant = {
        "projets": sorted((p.epic_trigramme, p.nom, p.date_fin) for p in source.query(Project)),
        "taches": sorted((t.nom, t.date_debut, t.date_fin) for t in source.query(Task)),
    }
    source.close()

    assert rapport.projets == 2
    assert rapport.taches == 3

    # Base vierge : seul l'admin qui exécute l'import préexiste.
    cible = session_factory()
    for p in cible.query(Project):
        cible.delete(p)
    cible.commit()
    assert cible.query(Project).count() == 0

    bilan = _relire(contenu, cible)
    assert not bilan.refus, bilan.refus

    apres = {
        "projets": sorted((p.epic_trigramme, p.nom, p.date_fin) for p in cible.query(Project)),
        "taches": sorted((t.nom, t.date_debut, t.date_fin) for t in cible.query(Task)),
    }
    cible.close()
    assert apres == avant


def test_la_date_de_debut_des_projets_est_perdue(client, auth, planning, session_factory):
    """Perte connue, épinglée ici plutôt que passée sous silence.

    La source ne porte pas de date de début de projet : l'import la fixe à
    `DEFAULT_PROJECT_START`. Le test précédent l'exclut donc de sa comparaison —
    ce qui la rendrait invisible. On l'affirme ici, pour que la perte reste un
    fait connu et qu'un futur ajout de colonne fasse échouer ce test au lieu de
    passer inaperçu.
    """
    from app.services.import_xlsx import DEFAULT_PROJECT_START

    db = session_factory()
    contenu, _ = exporter_classeur(db)
    debuts_avant = {p.date_debut for p in db.query(Project)}
    for p in db.query(Project):
        db.delete(p)
    db.commit()
    _relire(contenu, db)
    debuts_apres = {p.date_debut for p in db.query(Project)}
    db.close()

    assert debuts_avant != {DEFAULT_PROJECT_START}, "la fixture doit poser d'autres dates"
    assert debuts_apres == {DEFAULT_PROJECT_START}


def test_les_responsables_suivent_leur_tache(client, auth, fabrique, session_factory):
    """Rapprochés par NOM à l'import : c'est la seule prise que donne le format."""
    annuaire = client.get("/api/users/annuaire", headers=auth).json()
    fabrique.epic("ABC")
    p = fabrique.projet("ABC")
    fabrique.tache(p["id"], nom="Avec responsable", responsable_id=annuaire[0]["id"])

    db = session_factory()
    contenu, _ = exporter_classeur(db)
    db.close()

    feuille = load_workbook_sheet(contenu, "Detail des tache de projet")
    ligne = next(r for r in feuille if r[2] == "Avec responsable")
    assert ligne[6] == annuaire[0]["nom"], "le nom du responsable doit être écrit"


def test_rejouer_un_export_ne_duplique_rien(client, auth, planning, session_factory):
    """Deux imports du même export : le second ne crée rien.

    Les trigrammes sont FABRIQUÉS à chaque export ; si l'import s'y fiait pour
    reconnaître un projet, rejouer produirait des doublons. Il se fie à
    `(epic, nom)` — ce test verrouille ce point.
    """
    db = session_factory()
    contenu, _ = exporter_classeur(db)
    db.close()

    cible = session_factory()
    _relire(contenu, cible)
    total = cible.query(Project).count()
    bilan = _relire(contenu, cible)
    assert cible.query(Project).count() == total
    assert bilan.projets_crees == 0
    cible.close()


# --- ce que le format ne sait pas porter, et qui doit être DIT ---------------


def test_un_compte_desactive_n_est_pas_reexporte(client, auth, session_factory):
    """Le réémettre le ferait revenir ACTIF : l'import pose `actif: True` en dur.

    Une révocation annulée en silence par une réimportation serait un défaut de
    sécurité, pas une perte de données.
    """
    db = session_factory()
    parti = User(nom="Compte révoqué", email="parti@test.local", actif=False)
    db.add(parti)
    db.commit()
    contenu, rapport = exporter_classeur(db)
    db.close()

    noms = [r[0] for r in load_workbook_sheet(contenu, "Chargés de projets")]
    assert "Compte révoqué" not in noms


def test_les_jalons_signalent_la_perte_de_leur_rattachement(
    client, auth, fabrique, session_factory
):
    """L'import les regroupera sous « Jalons transverses » : autant l'annoncer."""
    fabrique.epic("ABC")
    p = fabrique.projet("ABC")
    fabrique.jalon([p["id"]], nom="Contrôle annuel")

    db = session_factory()
    _, rapport = exporter_classeur(db)
    db.close()

    assert rapport.jalons == 1
    assert any("rattachement" in m for m in rapport.omissions), rapport.omissions


def test_les_donnees_hors_format_sont_annoncees(client, auth, fabrique, session_factory):
    """Équipes, dépendances, mesures : aucun onglet ne les porte."""
    fabrique.equipe()

    db = session_factory()
    _, rapport = exporter_classeur(db)
    db.close()

    assert any("équipes" in m for m in rapport.omissions), rapport.omissions


def test_aucune_omission_annoncee_quand_tout_passe(client, auth, planning, session_factory):
    """L'inverse compte autant : un avertissement permanent ne serait plus lu."""
    db = session_factory()
    _, rapport = exporter_classeur(db)
    db.close()
    assert rapport.omissions == []


def test_une_tache_de_projet_non_planifie_est_signalee(client, auth, session_factory):
    """L'onglet des non planifiés ne porte pas de trigramme : rien à rattacher.

    La tâche ne peut pas être exportée. Ce qu'on refuse d'écrire doit se dire —
    une ligne disparue en silence est ce qu'un export ne doit jamais faire.
    """
    r = client.post(
        "/api/epics",
        json={"trigramme": "NPL", "nom": "Non planifiés", "statut": "idee",
              "categorie": "operationnel"},
        headers=auth,
    )
    assert r.status_code == 201, r.text
    p = client.post(
        "/api/projects",
        json={"epic_trigramme": "NPL", "nom": "À classer",
              "date_debut": "2026-05-01", "date_fin": "2028-12-31"},
        headers=auth,
    ).json()
    client.post(
        "/api/tasks",
        json={"projet_id": p["id"], "nom": "Orpheline",
              "date_debut": "2026-06-01", "date_fin": "2026-06-10"},
        headers=auth,
    )

    db = session_factory()
    _, rapport = exporter_classeur(db)
    db.close()

    assert rapport.taches == 0
    assert rapport.projets_non_planifies == 1
    assert any("Orpheline" in m for m in rapport.omissions), rapport.omissions


# --- utilitaire ---------------------------------------------------------------


def load_workbook_sheet(contenu: bytes, onglet: str):
    """Lignes de données d'un onglet (en-tête sautée)."""
    from io import BytesIO

    wb = load_workbook(BytesIO(contenu))
    return list(wb[onglet].iter_rows(values_only=True))[1:]

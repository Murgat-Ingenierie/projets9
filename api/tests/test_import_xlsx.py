"""Import du classeur depuis l'application (étape 4).

Ce qui doit être prouvé ici n'est pas « le fichier est lu », mais que l'import
**passe toujours par les invariants**. C'était la propriété du script CLI, qui
écrivait via HTTP : la déplacer dans l'API ne doit pas la perdre. D'où le test
central, `test_une_ligne_invalide_est_refusee_et_rapportee` : une ligne qui viole
un invariant doit être **refusée ET signalée**, jamais insérée en douce.

Les classeurs sont fabriqués ici même, avec les onglets et indices de colonnes du
format réel (cf. docstring de `app/services/import_xlsx.py`).
"""

from io import BytesIO

import pytest
from openpyxl import Workbook


def classeur(projets=(), taches=(), charges=("Alice Martin",)) -> bytes:
    """Fabrique un classeur au format source. Colonnes = indices réels."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Chargés de projets"
    ws.append(["Nom"])
    for c in charges:
        ws.append([c])

    p = wb.create_sheet("Projets")
    p.append(["Nom", "fin prévu", "Jalon de fin maximum", "Raison", "Trigramme",
              "Epic lié", "Rappel", "Terminé"])
    for ligne in projets:
        p.append(list(ligne))

    t = wb.create_sheet("Detail des tache de projet")
    t.append(["Projet lié", "Rappel", "Nom de la tache", "Date de début",
              "Jalon maximum", "Durée", "Responsable", "Equipe", "Materiel", "Terminé"])
    for ligne in taches:
        t.append(list(ligne))

    tampon = BytesIO()
    wb.save(tampon)
    return tampon.getvalue()


def envoi(client, auth, contenu: bytes, nom: str = "source.xlsx"):
    return client.post(
        "/api/import/xlsx",
        files={"fichier": (nom, contenu,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth,
    )


# --- le cas qui compte -------------------------------------------------------


def test_une_ligne_invalide_est_refusee_et_rapportee(client, auth):
    """La propriété héritée du script : les invariants s'appliquent toujours.

    On reproduit le cas RÉEL rencontré au premier import : un projet dont
    l'échéance dépasse la date de fin prévue de son epic (INV-10). Neuf projets
    avaient été refusés ainsi — et c'est ce refus qui avait révélé une
    incohérence dans la source. Il doit donc remonter à l'utilisateur.

    (Un `date_fin` aberrant ne conviendrait pas pour ce test : l'import
    normalise déjà les dates antérieures au début par défaut.)
    """
    r = client.post(
        "/api/epics",
        json={"trigramme": "BOR", "nom": "Epic borné", "statut": "idee",
              "categorie": "operationnel", "date_fin_prevue": "2026-12-31"},
        headers=auth,
    )
    assert r.status_code == 201, r.text

    contenu = classeur(projets=[
        ["Projet tenable", None, 2026, "", "OK1", "BOR", "", None],
        ["Projet trop long", None, 2028, "", "KO1", "BOR", "", None],
    ])
    r = envoi(client, auth, contenu)
    assert r.status_code == 200, r.text
    rapport = r.json()

    assert rapport["projets_crees"] >= 1, "le projet tenable doit passer"
    assert rapport["refus"], "un refus doit être RAPPORTÉ, pas avalé"
    assert any("INV-10" in m for m in rapport["refus"]), rapport["refus"]

    noms = [p["nom"] for p in client.get("/api/projects", headers=auth).json()]
    assert "Projet trop long" not in noms, "rien ne doit être inséré en douce"
    assert "Projet tenable" in noms


def test_import_idempotent(client, auth):
    """Rejouer le même classeur ne duplique rien — le script le garantissait."""
    contenu = classeur(projets=[["Bassin nord", None, 2027, "", "BN1", "AAA", "Epic A", None]])
    assert envoi(client, auth, contenu).status_code == 200
    premier = len(client.get("/api/projects", headers=auth).json())

    r2 = envoi(client, auth, contenu)
    assert r2.status_code == 200
    assert r2.json()["projets_deja_presents"] >= 1
    assert len(client.get("/api/projects", headers=auth).json()) == premier


def test_les_taches_sont_rattachees_a_leur_projet(client, auth):
    contenu = classeur(
        projets=[["Bassin sud", None, 2027, "", "BS1", "AAA", "Epic A", None]],
        taches=[["BS1", "", "Poser les capteurs", "2026-06-01", None, 10, "Alice Martin", "", "", None]],
    )
    r = envoi(client, auth, contenu)
    assert r.status_code == 200, r.text
    assert r.json()["taches_creees"] == 1
    assert r.json()["taches_sans_projet"] == 0


def test_tache_sans_projet_connu_est_comptee(client, auth):
    """Ne pas la perdre en silence : c'est ce qui avait révélé 34 orphelines."""
    contenu = classeur(taches=[
        ["INEXISTANT", "", "Tâche orpheline", "2026-06-01", None, 5, "Alice Martin", "", "", None]
    ])
    r = envoi(client, auth, contenu)
    assert r.status_code == 200
    assert r.json()["taches_sans_projet"] == 1


# --- entrées invalides : dire lesquelles, plutôt que planter -----------------


def test_fichier_non_xlsx_refuse(client, auth):
    r = envoi(client, auth, b"ceci n'est pas un classeur", nom="donnees.csv")
    assert r.status_code == 400
    assert "xlsx" in r.json()["detail"].lower()


def test_classeur_sans_les_onglets_attendus(client, auth):
    """Cas courant : le mauvais fichier. Le message doit être actionnable."""
    wb = Workbook()
    wb.active.title = "Feuille1"
    tampon = BytesIO()
    wb.save(tampon)

    r = envoi(client, auth, tampon.getvalue())
    assert r.status_code == 400
    detail = r.json()["detail"]
    assert "manquant" in detail.lower()
    assert "Feuille1" in detail, "doit dire ce qu'il a trouvé, pas seulement ce qui manque"


def test_fichier_vide_refuse(client, auth):
    assert envoi(client, auth, b"").status_code == 400


# --- autorisation -------------------------------------------------------------


def test_import_reserve_aux_administrateurs(client, auth_membre):
    """Un import touche tout le planning : ce n'est pas une action de membre."""
    contenu = classeur(projets=[["X", None, 2027, "", "XX1", "AAA", "E", None]])
    assert envoi(client, auth_membre, contenu).status_code == 403


@pytest.mark.parametrize("ressource", ["epics", "projects", "tasks", "users"])
def test_le_rapport_donne_les_totaux(client, auth, ressource):
    contenu = classeur(projets=[["Total", None, 2027, "", "TO1", "AAA", "Epic A", None]])
    totaux = envoi(client, auth, contenu).json()["totaux"]
    assert ressource in totaux
